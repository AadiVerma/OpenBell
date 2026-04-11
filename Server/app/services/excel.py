"""
Excel report generation (openpyxl).
Dark-themed workbook with four sheets:
  Summary · Buy Signals · All Signals · Bearish Watch
"""
from __future__ import annotations

import datetime
import io
import math

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import Settings
from app.models.prediction import Prediction

# ── Colour palette ────────────────────────────────────────────────────────────
_BG_DARK = "1A1A2E"
_BG_MID = "16213E"
_BG_CARD = "0F3460"
_GREEN = "00B86B"
_RED = "FF4757"
_AMBER = "FFA502"
_WHITE = "FFFFFF"
_GREY = "8892B0"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 11, color: str = _WHITE) -> Font:
    return Font(bold=bold, size=size, color=color, name="Calibri")


def _position_alloc(confidence: int, budget: float) -> float:
    if confidence >= 80:
        return round(budget * 0.15, 2)
    if confidence >= 70:
        return round(budget * 0.10, 2)
    if confidence >= 60:
        return round(budget * 0.07, 2)
    return round(budget * 0.05, 2)


def generate_excel(
    predictions: list[Prediction],
    report_date: datetime.date,
    settings: Settings,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _summary_sheet(wb, predictions, report_date, settings)
    _buy_signals_sheet(wb, predictions, settings)
    _all_signals_sheet(wb, predictions)
    _bearish_sheet(wb, predictions)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _summary_sheet(
    wb: openpyxl.Workbook,
    predictions: list[Prediction],
    report_date: datetime.date,
    settings: Settings,
) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30

    bullish = [p for p in predictions if p.signal == "bullish"]
    bearish = [p for p in predictions if p.signal == "bearish"]
    neutral = [p for p in predictions if p.signal == "neutral"]
    avg_conf = (
        round(sum(p.confidence for p in predictions) / len(predictions))
        if predictions
        else 0
    )

    rows = [
        ("OpenBell Report", report_date.strftime("%d %B %Y")),
        ("Total Stocks", len(predictions)),
        ("Bullish Signals", len(bullish)),
        ("Bearish Signals", len(bearish)),
        ("Neutral Signals", len(neutral)),
        ("Avg Confidence", f"{avg_conf}%"),
        ("Portfolio Budget", f"₹{settings.PORTFOLIO_BUDGET:,.2f}"),
        ("", ""),
        ("Disclaimer", "AI-generated signals. Not financial advice."),
    ]
    for i, (label, value) in enumerate(rows, 1):
        ws.cell(i, 1, label).fill = _fill(_BG_MID)
        ws.cell(i, 1).font = _font(bold=True, color=_GREY)
        ws.cell(i, 2, str(value)).fill = _fill(_BG_DARK)
        ws.cell(i, 2).font = _font()


def _buy_signals_sheet(
    wb: openpyxl.Workbook,
    predictions: list[Prediction],
    settings: Settings,
) -> None:
    ws = wb.create_sheet("Buy Signals")
    ws.sheet_view.showGridLines = False

    headers = [
        "Ticker", "Conf%", "Current", "Limit Entry",
        "Target Low", "Target High", "Upside%",
        "Stop Loss", "Alloc ₹", "Qty", "Max Profit", "Max Loss",
    ]
    widths = [12, 8, 12, 12, 12, 12, 10, 12, 12, 8, 12, 12]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, col, h)
        cell.fill = _fill(_BG_CARD)
        cell.font = _font(bold=True, color=_GREEN)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    bullish = sorted(
        [p for p in predictions if p.signal == "bullish" and p.confidence >= 60],
        key=lambda p: p.confidence,
        reverse=True,
    )
    for row, p in enumerate(bullish, 2):
        alloc = _position_alloc(p.confidence, settings.PORTFOLIO_BUDGET)
        qty = max(1, math.floor(alloc / p.limit_price)) if p.limit_price else 0
        stop = round(p.limit_price * 0.97, 2)
        upside = (
            round((p.target_high - p.limit_price) / p.limit_price * 100, 2)
            if p.limit_price
            else 0
        )
        max_profit = round((p.target_high - p.limit_price) * qty, 2)
        max_loss = round((p.limit_price - stop) * qty, 2)

        bg = _fill(_BG_DARK if row % 2 == 0 else _BG_MID)
        for col, val in enumerate(
            [
                p.ticker, f"{p.confidence}%", p.current_price, p.limit_price,
                p.target_low, p.target_high, f"{upside}%",
                stop, alloc, qty, max_profit, max_loss,
            ],
            1,
        ):
            cell = ws.cell(row, col, val)
            cell.fill = bg
            cell.font = _font()
            cell.alignment = Alignment(horizontal="center")


def _all_signals_sheet(wb: openpyxl.Workbook, predictions: list[Prediction]) -> None:
    ws = wb.create_sheet("All Signals")
    ws.sheet_view.showGridLines = False

    headers = ["Ticker", "Signal", "Conf%", "Direction", "Target Low", "Target High", "Limit", "Reasoning"]
    widths = [12, 10, 8, 10, 12, 12, 12, 65]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, col, h)
        cell.fill = _fill(_BG_CARD)
        cell.font = _font(bold=True)
        ws.column_dimensions[cell.column_letter].width = w

    for row, p in enumerate(
        sorted(predictions, key=lambda p: p.confidence, reverse=True), 2
    ):
        sig_color = _GREEN if p.signal == "bullish" else (_RED if p.signal == "bearish" else _AMBER)
        bg = _fill(_BG_DARK if row % 2 == 0 else _BG_MID)
        vals = [p.ticker, p.signal.upper(), f"{p.confidence}%", p.predicted_direction,
                p.target_low, p.target_high, p.limit_price, p.reasoning]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.fill = bg
            cell.font = _font(color=sig_color if col == 2 else _WHITE)
            if col == len(vals):
                cell.alignment = Alignment(wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 40


def _bearish_sheet(wb: openpyxl.Workbook, predictions: list[Prediction]) -> None:
    ws = wb.create_sheet("Bearish Watch")
    ws.sheet_view.showGridLines = False

    headers = ["Ticker", "Conf%", "Current", "Target Low", "Downside%", "Reasoning"]
    widths = [12, 8, 12, 12, 12, 65]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, col, h)
        cell.fill = _fill(_BG_CARD)
        cell.font = _font(bold=True, color=_RED)
        ws.column_dimensions[cell.column_letter].width = w

    bearish = sorted(
        [p for p in predictions if p.signal == "bearish" and p.confidence >= 60],
        key=lambda p: p.confidence,
        reverse=True,
    )
    for row, p in enumerate(bearish, 2):
        downside = (
            round((p.target_low - p.current_price) / p.current_price * 100, 2)
            if p.current_price
            else 0
        )
        bg = _fill(_BG_DARK if row % 2 == 0 else _BG_MID)
        for col, val in enumerate(
            [p.ticker, f"{p.confidence}%", p.current_price, p.target_low, f"{downside}%", p.reasoning],
            1,
        ):
            cell = ws.cell(row, col, val)
            cell.fill = bg
            cell.font = _font()
            if col == 6:
                cell.alignment = Alignment(wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 40
